
# [START app]
import logging

# [START imports]
#from flask import Flask, render_template, request
#import flask
#from flask_bootstrap import Bootstrap
#from flask_wtf import FlaskForm
#from wtforms import StringField
#from wtforms.validators import DataRequired
import openpyxl
from openpyxl import Workbook, load_workbook
#import numpy as np
#import os
import csv
import time
import numpy as np
import sys
# [END imports]




class NaiveBayesNetworkNode:

    # Initializer
    def __init__(self, name, states):
        #self.name = name.replace('_', ' ')
        self.name = name
        self.states = states
        self.parent = ''
        self.category = ''
        self.menuStates = []
        self.menuStates.append(self.name)
        for s in self.states:
            self.menuStates.append(self.name+":"+s)
        self.value = ''
        self.menuValue = ''
        self.priors = np.ones(len(self.states))
        self.priors = self.priors / len(self.priors)
        self.probs = []
        self.hasChoice = ''
        self.sensitive = ''
        self.style = ''
        self.modified = time.time()


    def set_parent(self, node):
        self.parent = node.name

    def is_state_selected(self, menuChoice):
        if menuChoice == self.menuValue:
            return 'selected'
        return ''



class NaiveBayesNetwork:

    def __init__(self):
        self.filename = ""
        self.name = "ExampleNetwork"
        self.categories = ["Signal", "Spatial", "Time", "Miscellaneous", "Clinical"];
        self.nodes = []
        self.nodeMap = {}
        self.modified = time.time()

    # Set all node states to clear
    def reset(self):
        self.modified = time.time()
        for n in self.nodes:
            self.clear_node_state(n.name)

    def add_node(self, node):
        self.modified = time.time()
        self.nodes.append(node)
        self.nodeMap[node.name] = len(self.nodes)-1
        #print( node.name + str( len(self.nodes)-1 ))

    def has_node(self,name):
        return self.nodeMap.has_key(name)

    def get_node(self, name):
        if self.has_node(name):
            return self.nodes[ self.nodeMap[name] ]
        return None

    def get_node_states(self, name):
        if self.has_node(name):
            return self.nodes[ self.nodeMap[name] ].states
        return None

    def set_node_state(self, name, state):
        if self.has_node(name):
            self.modified = time.time()
            n = self.nodeMap[name]
            if state in self.nodes[n].states:
                self.nodes[n].value = state
                self.nodes[n].menuValue = self.nodes[n].name + ':' + state
                self.nodes[n].hasChoice = 'hasChoice'
                self.nodes[n].sensitive = ''
            else:
                raise Exception("Invalid state set: " + state )

    def set_node_states_by_result(self, index):
        self.modified = time.time()
        for n in self.nodes:
            if n.name != 'Diagnosis':
                vals = n.probs[index,:].tolist()
                state = vals.index(max(vals))
                self.set_node_state(n.name, n.states[state])

    def clear_node_state(self, name):
        if self.has_node(name):
            self.modified = time.time()
            n = self.nodeMap[name]
            self.nodes[n].value = ''
            self.nodes[n].menuValue = ''
            self.nodes[n].hasChoice = ''
        else:
            raise Exception("Node does not exist: " + name )

    def set_node_probs(self, name, probs):
        if self.has_node(name):
            self.modified = time.time()
            n = self.nodeMap[name]
            self.nodes[n].probs = probs
        else:
            raise Exception("Node does not exist: " + name )

    def get_node_priors(self, name):
        if self.has_node(name):
            return( self.nodes[ self.nodeMap[name] ].priors)
        return None

    def get_node_category(self, name):
        if self.has_node(name):
            return( self.nodes[ self.nodeMap[name] ].category)
        return None

    def number_of_nodes_in_category(self,category):
        count = 0
        for n in self.nodes:
            if n.category == category:
                count = count + 1
        return count

    def names_of_nodes_in_category(self,category):
        names = []
        for n in self.nodes:
            if n.category == category:
                names.append(n.name)
        return names

    def nodes_in_category(self,category):
        nodes = []
        for n in self.nodes:
            if n.category == category:
                nodes.append(n)
        return nodes

    def read_csv(self, network_file):
        self.filename = network_file
        self.modified = time.time()

        diseases = []
        diseasePrior = []
        nodeNames = []
        nodeCategories = []
        nodeStates = []

        self.categories = []
        prob = []

        with open(self.filename) as csv_file:
            csv_reader =  csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if  line_count == 0:
                    allNodeStates = row[2:len(row)]

                    for n in allNodeStates:
                        nodeParts = n.split(":")
                        nodeCategories.append(nodeParts[0])
                        nodeNames.append(nodeParts[1])
                        nodeStates.append(nodeParts[2])

                    for n in np.unique(nodeNames):
                        states = []
                        cat = ''
                        for i in range(0,len(nodeNames)):
                            if nodeNames[i] == n:
                                states.append(nodeStates[i])
                                cat = nodeCategories[i]
                        node = NaiveBayesNetworkNode(n,states)
                        node.category = cat
                        node.parent = "Diagnosis"
                        self.add_node(node)

                else:
                    diseases.append(row[0])
                    diseasePrior.append(row[1])
                    rValues = row[2:]
                    prob.append(rValues)


                line_count += 1

        prob = np.vstack(prob)

        for n in self.nodes:
            nMat = np.empty((len(diseases), len(n.states) ))
            for s in range(0, len(n.states)):
                for i in range(0,len(nodeNames)):
                    if nodeNames[i] == n.name and n.states[s] == nodeStates[i]:
                        nMat[:,s] = prob[:,i]

            self.set_node_probs( n.name, nMat )

        dNode = NaiveBayesNetworkNode("Diagnosis", diseases)
        dNode.category = "Diagnosis"
        dNode.priors = diseasePrior
        self.add_node(dNode)
        self.categories = np.unique(nodeCategories)

    def calculate_node_sensitivity(self, name, currentProbs):

        if ( not(self.has_node(name)) ):
            return 0

        node = self.nodes[ self.nodeMap[name] ]
        val = node.value
        probs = self.get_node_priors("Diagnosis")*0

        for v in range(len(node.states)):
            self.set_node_state(node.name, node.states[v])
            s, vProbs = self.get_diagnoses(False)
            probs += vProbs / ( currentProbs * node.probs[:,v] )
            probs[np.isnan(probs)] = 0

        if val == '':
            self.clear_node_state(node.name)
        else:
            self.set_node_state(node.name, val)

        sens = probs.sum()
        return sens


    def get_diagnoses(self, isRadiologic):
        mat = []
        if not(isRadiologic):
            mat.append(self.get_node_priors("Diagnosis"))
        else:
            unityPrior = []
            dis = self.get_node_priors("Diagnosis")
            for x in dis:
                unityPrior.append(1.0/len(dis))
            mat.append(unityPrior)

        for n in self.nodes:
            if (n.value != ''):
                if not(isRadiologic):
                    idx = n.states.index(n.value)
                    #print(n.name + ' ' + str(idx) + ' ' + n.value)
                    #print(n.probs)
                    #print( n.probs[:,idx])
                    mat.append( n.probs[:,idx])
                elif (n.category != "Clinical"):
                    #print(n.category + n.name)
                    idx = n.states.index(n.value)
                    mat.append( n.probs[:,idx])


        mat = np.vstack(mat)
        mat = mat.astype('double')
        mat = np.prod(mat, axis=0)
        mat = mat/np.sum(mat)
        sorted = np.argsort(mat)
        sorted = sorted[::-1]
        return sorted, mat


    def read_sheet(self, table_file, tableName):
    #    tableDir = "/home/jiancong/Desktop/projects/BayesNet/NaiveBayes/BasalGanglia/BG_Bayesian_network_NaiveBayes.xlsx"
    #    tableName = "BGnetwork"
        self.filename = table_file
        self.name = tableName
        self.modified = time.time()

        wb=openpyxl.load_workbook(table_file)
        worksheet = wb[tableName]

        #Read the keys
        rows = list(worksheet.rows)
        first_row = rows[0]
        second_row = rows[1]

        keys = [c.value for c in first_row]
        values = [c.value for c in second_row]

        values = [v for v in values if v is not None]
        keys = keys[:len(values)]

        values = values[1:]
        keys = keys[1:]

        key_value_dict = {}

        current_k = ""
        for k, v, i in zip(keys, values, range(2, len(values) + 2)):
            if k is not None:
                current_k = k.split(" ")[0].upper()
                key_value_dict[current_k] = {v.split(" ")[0].upper():i - 2}
            else:
                key_value_dict[current_k][v.split(" ")[0].upper()] = i - 2

        # Read the probabilty
        disease = []
        prob = []
        for row in rows[2:]:
            row_values = [c.value for c in row if c.value is not None]

            if len(row_values) ==0:
                break

            if row_values[0] is None:
                break
            disease.append(row_values[0])
            prob.append(row_values[1:])
        prob = np.array(prob)
        prob = prob/100.0

        return disease, prob, key_value_dict

    def sheet_to_class(self, table_file, tableName, classFileName, className):
    #    tableDir = "/home/jiancong/Desktop/projects/BayesNet/NaiveBayes/BasalGanglia/BG_Bayesian_network_NaiveBayes.xlsx"
    #    tableName = "BGnetwork"

        file = open(classFileName,"w+")

        self.filename = table_file
        self.name = tableName
        self.modified = time.time()

        wb=openpyxl.load_workbook(table_file)
        worksheet = wb[tableName]


        #Read the keys
        rows = list(worksheet.rows)
        first_row = rows[0]
        second_row = rows[1]

        # Read the diagnoses and priors
        diagnoses = []
        prior = []
        probTable = []
        for row in rows[2:]:
            row_values = [c.value for c in row if c.value is not None]

            if len(row_values) ==0:
                break

            if row_values[0] is None:
                break
            diagnoses.append(row_values[0])
            prior.append(row_values[1])
            probTable.append(row_values[2:])
        probTable = np.array(probTable)/100.0

        prior = np.array(prior)
        prior = prior/100.0
        states = []
        for d in diagnoses:
            dx = "\'" + d + "\'"
            states.append(dx)
        states = ",".join(states)
        priorString = "["
        for i in range(len(prior) ):
            if ( i > 0 ):
                priorString += ","
            priorString += str(prior[i])
        priorString += "]"

        print(probTable)




        file.write( "from easybayesy import NaiveBayesNetworkNode, NaiveBayesNetwork\n")
        file.write( "import numpy as np\n")
        file.write( "import openpyxl\n")
        file.write( "from openpyxl import Workbook, load_workbook\n\n")

        file.write( "def " + className + "(): \n ")
        file.write( "  network =  NaiveBayesNetwork(); \n\n" )

        file.write( "   dx = NaiveBayesNetworkNode(\'Diagnosis\'" + ",[" + states + "]) \n" )
        file.write( "   dx.category=\'Diagnosis\' \n" )
        file.write( "   nDx = " + str(len(diagnoses)) + "\n" )
        file.write( "   dx.priors = np.array( " + priorString + ") \n" )

        file.write( "   network.add_node(dx) \n\n")

        nDx = len(diagnoses)


        keys = [c.value for c in first_row]
        values = [c.value for c in second_row]

        values = [v for v in values if v is not None]
        keys = keys[:len(values)]

        values = values[2:]
        keys = keys[2:]

        key_value_dict = {}

        current_k = ""
        for k, v, i in zip(keys, values, range(2, len(values) + 2)):
            if k is not None:
                current_k = k.split(" ")[0]
                print(current_k)
                key_value_dict[current_k] = {v.split(" ")[0]:i - 2}

            else:
                key_value_dict[current_k][v.split(" ")[0]] = i - 2


        categories = []
        for k in key_value_dict.keys():
            vals = key_value_dict[k]
            print(k)
            print(vals.values())
            kProbs = []
            for v in vals.values():
                kProbs.append( probTable[:,v])
            kProbs = np.array(kProbs)
            print(kProbs)

            kProbString = "[ "
            for i in range(kProbs.shape[0]):
                kProbString += "[ "
                for j in range(kProbs.shape[1]):
                    p = kProbs[i,j]
                    kProbString += str(p)
                    if ( j < kProbs.shape[1]-1 ):
                        kProbString += ","
                    else:
                        kProbString += " ]"
                if ( i < kProbs.shape[0]-1 ):
                    kProbString += ", "
                else:
                    kProbString += " ]"

            nodeName = k.split(":")[1]
            nodeCat = k.split(":")[0]
            categories.append(nodeCat)

            states = []
            probs = np.array( (len(vals), nDx ))
            for v in vals:
                state = "\'" + str(v) + "\'"
                states.append(state)

            states = ",".join(states)

            file.write( "   n = NaiveBayesNetworkNode('" + nodeName + "'" + ",[" + states + "]) \n" )
            file.write( "   n.category = '" + nodeCat + "'\n");
            file.write( "   n.parent = \'Diagnosis\' \n" )
            file.write( "   n.probs = np.array( " + kProbString + ").transpose() \n")
            file.write( "   network.add_node(n) \n\n")

        file.write( "   network.categories = ['Signal', 'Spatial', 'Clinical'] \n" )
        file.write( "   return network\n");

        file.close()
        return 0

        #return disease, prob, key_value_dict
